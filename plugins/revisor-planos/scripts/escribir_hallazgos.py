"""Escribe hallazgos hidrosanitarios en el Excel del proyecto (append).

Lee un JSON (lista de objetos con las 10 claves del esquema + 'plano_referencia')
y agrega una fila por hallazgo en la hoja 'Hallazgos hidrosanitarios' del Excel
destino, sin sobrescribir filas existentes.

Si el Excel destino no existe, se crea automaticamente copiando la plantilla
templates/hallazgos_hidrosanitario.xlsx del plugin.

Mapeo de columnas del esquema v2 (encabezado en la fila 4, datos desde la fila 5):
    A  ID                     -> consecutivo, continua el maximo existente
    B  disciplina
    C  hallazgo
    D  ubicacion
    E  severidad              -> que tan grave es el hallazgo si es real
    F  nivel_confianza        -> Alta / Media / Baja (seguridad del agente)
    G  referencia_normativa
    H  riesgo
    I  justificacion_tecnica
    J  accion_correctiva
    K  plano_referencia
    L  fecha_deteccion        -> por defecto la fecha de hoy
    M  estado                 -> seguimiento MANUAL; siempre se escribe 'Pendiente'

Los archivos con el esquema v1 (con 'requiere_validacion' y 'estado' en K) se
rechazan antes de escribir; migralos con scripts/migrar_esquema.py.

Uso:
    python escribir_hallazgos.py <ruta_json> <ruta_xlsx> [--fecha AAAA-MM-DD] [--dry-run]

Codigos de salida:
    0  OK
    2  error de permisos / archivo en uso (mensaje accionable en stderr)
    3  error de entrada (JSON invalido, plantilla ausente, hoja o esquema incorrecto)
    4  falta una dependencia de Python
"""
import argparse
import datetime as dt
import json
import os
import shutil
import sys
import tempfile

SALIDA_OK = 0
SALIDA_PERMISOS = 2
SALIDA_ENTRADA = 3
SALIDA_DEPENDENCIA = 4

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - depende del entorno
    sys.stderr.write(
        "ERROR: falta la dependencia 'openpyxl'.\n"
        "Instalala con:\n"
        f'  "{sys.executable}" -m pip install openpyxl\n'
    )
    sys.exit(SALIDA_DEPENDENCIA)

HOJA = "Hallazgos hidrosanitarios"
FILA_ENCABEZADO = 4
FILA_INICIO_DATOS = 5
DIR = os.path.dirname(os.path.abspath(__file__))
RAIZ_PLUGIN = os.environ.get("CLAUDE_PLUGIN_ROOT") or os.path.dirname(DIR)
PLANTILLA = os.path.join(RAIZ_PLUGIN, "templates", "hallazgos_hidrosanitario.xlsx")

# Encabezado esperado en la fila 4, en orden (columna = indice + 1).
ESQUEMA = [
    "ID",
    "disciplina",
    "hallazgo",
    "ubicacion",
    "severidad",
    "nivel_confianza",
    "referencia_normativa",
    "riesgo",
    "justificacion_tecnica",
    "accion_correctiva",
    "plano_referencia",
    "fecha_deteccion",
    "estado",
]

# Esquema anterior, solo para reconocerlo y dar un mensaje util.
ESQUEMA_V1 = [
    "ID",
    "disciplina",
    "hallazgo",
    "ubicacion",
    "severidad",
    "referencia_normativa",
    "riesgo",
    "justificacion_tecnica",
    "accion_correctiva",
    "requiere_validacion",
    "estado",
    "plano_referencia",
    "fecha_deteccion",
]

# columna (indice 1-based) -> clave del JSON que aporta el agente
CAMPOS = {
    2: "disciplina",
    3: "hallazgo",
    4: "ubicacion",
    5: "severidad",
    6: "nivel_confianza",
    7: "referencia_normativa",
    8: "riesgo",
    9: "justificacion_tecnica",
    10: "accion_correctiva",
}

COL_PLANO = 11
COL_FECHA = 12
COL_ESTADO = 13

ESTADO_INICIAL = "Pendiente"
NIVELES_CONFIANZA = ("Alta", "Media", "Baja")


class ErrorEntrada(Exception):
    """Problema con los datos o rutas de entrada; mensaje ya legible."""


def normalizar_confianza(valor):
    """'  alta ' -> 'Alta'. Devuelve None si no es un nivel valido.

    Se toleran mayusculas/minusculas y espacios sobrantes; cualquier otro valor
    (incluido vacio o ausente) se rechaza, nunca se adivina.
    """
    if valor is None:
        return None
    texto = str(valor).strip()
    for nivel in NIVELES_CONFIANZA:
        if texto.lower() == nivel.lower():
            return nivel
    return None


def excel_abierto(ruta):
    """Detecta el lockfile '~$archivo.xlsx' que Excel crea al abrir un libro."""
    carpeta, nombre = os.path.split(os.path.abspath(ruta))
    return os.path.exists(os.path.join(carpeta, "~$" + nombre))


def mensaje_permisos(ruta, err):
    """Mensaje accionable para un fallo de escritura en Windows."""
    lineas = [
        "ERROR DE PERMISOS: no se pudo escribir el archivo.",
        f"  Ruta: {os.path.abspath(ruta)}",
        f"  Detalle del sistema: {err}",
        "",
    ]
    if excel_abierto(ruta):
        lineas += [
            "Causa mas probable: el archivo esta ABIERTO en Excel.",
            "Como solucionarlo:",
            "  1. Cierra el archivo en Excel (incluida cualquier ventana de solo lectura).",
            "  2. Vuelve a ejecutar el comando.",
        ]
    else:
        lineas += [
            "Causas posibles (en orden de probabilidad):",
            "  1. El archivo esta abierto en Excel -> cierralo y reintenta.",
            "  2. Acceso controlado a carpetas de Windows Defender esta bloqueando la",
            "     escritura (tipico en Escritorio, Documentos, Imagenes). Ve a",
            "     Seguridad de Windows -> Proteccion contra virus y amenazas ->",
            "     Proteccion contra ransomware -> Permitir una app a traves del acceso",
            "     controlado, y autoriza:",
            f"       {sys.executable}",
            "  3. La carpeta es de solo lectura o esta en OneDrive sin sincronizar ->",
            "     elige otra ruta destino, por ejemplo dentro de la carpeta del proyecto.",
        ]
    lineas += [
        "",
        "NO se escribio ningun hallazgo. El archivo destino quedo intacto.",
    ]
    return "\n".join(lineas) + "\n"


def asegurar_destino(ruta_xlsx, dry_run=False):
    """Crea el Excel destino desde la plantilla si aun no existe.

    Devuelve True si lo creo (o lo crearia, en dry-run).
    """
    if os.path.exists(ruta_xlsx):
        return False

    if not os.path.exists(PLANTILLA):
        raise ErrorEntrada(
            "No existe el Excel destino y tampoco se encontro la plantilla del plugin.\n"
            f"  Destino: {os.path.abspath(ruta_xlsx)}\n"
            f"  Plantilla esperada: {PLANTILLA}\n"
            "Verifica la instalacion del plugin o crea el Excel manualmente."
        )

    if dry_run:
        return True

    carpeta = os.path.dirname(os.path.abspath(ruta_xlsx))
    os.makedirs(carpeta, exist_ok=True)
    shutil.copyfile(PLANTILLA, ruta_xlsx)
    return True


def verificar_esquema(ws, ruta_xlsx):
    """Compara la fila 4 con el esquema v2 y aborta si no coincide.

    Sin esta comprobacion, un Excel del esquema v1 (mismo numero de columnas,
    distinto orden) aceptaria la escritura y dejaria cada dato en la columna
    equivocada, en silencio.
    """
    encontrado = [ws.cell(FILA_ENCABEZADO, c).value for c in range(1, len(ESQUEMA) + 1)]
    encontrado = [("" if v is None else str(v).strip()) for v in encontrado]
    if encontrado == ESQUEMA:
        return

    lineas = [
        "El Excel destino no usa el esquema de columnas actual (v2).",
        f"  Ruta: {os.path.abspath(ruta_xlsx)}",
        "",
    ]
    if encontrado == ESQUEMA_V1:
        lineas += [
            "Es un archivo del esquema ANTERIOR (v1): tiene 'requiere_validacion' y",
            "'estado' en la columna K. Escribir sobre el dejaria cada dato en la",
            "columna equivocada, asi que no se escribio nada.",
            "",
            "Migralo conservando los hallazgos ya registrados:",
            f'  python "{os.path.join(RAIZ_PLUGIN, "scripts", "migrar_esquema.py")}" '
            f'"{os.path.abspath(ruta_xlsx)}"',
        ]
    else:
        diferencias = [
            f"    columna {chr(64 + i)}: se esperaba '{esp}' y hay '{hay}'"
            for i, (esp, hay) in enumerate(zip(ESQUEMA, encontrado), start=1)
            if esp != hay
        ]
        lineas += ["Diferencias en la fila 4:"] + diferencias + [
            "",
            "Usa un Excel generado desde la plantilla del plugin, o corrige el",
            "encabezado para que coincida exactamente con el esquema.",
        ]
    lineas += ["", "NO se escribio ningun hallazgo. El archivo destino quedo intacto."]
    raise ErrorEntrada("\n".join(lineas))


def validar_hallazgos(hallazgos):
    """Normaliza 'nivel_confianza' in-place; acumula todos los errores y aborta.

    Se valida antes de tocar el disco: si algo esta mal, no se crea ni se
    modifica ningun archivo.
    """
    errores = []
    for i, h in enumerate(hallazgos, start=1):
        if not isinstance(h, dict):
            errores.append(f"  hallazgo #{i}: se esperaba un objeto JSON, hay {type(h).__name__}")
            continue
        nivel = normalizar_confianza(h.get("nivel_confianza"))
        if nivel is None:
            crudo = h.get("nivel_confianza")
            detalle = "falta la clave" if crudo is None else f"valor invalido: {crudo!r}"
            errores.append(f"  hallazgo #{i}: 'nivel_confianza' {detalle}")
        else:
            h["nivel_confianza"] = nivel

    if errores:
        raise ErrorEntrada(
            "Hay hallazgos que no cumplen el esquema.\n"
            + "\n".join(errores)
            + "\n\n'nivel_confianza' es obligatorio y debe ser exactamente uno de: "
            + ", ".join(NIVELES_CONFIANZA)
            + ".\n(se aceptan mayusculas/minusculas y espacios sobrantes)\n"
            "  Alta  = evidencia directa y clara en el plano\n"
            "  Media = evidencia parcial o depende de contexto adicional\n"
            "  Baja  = interpretacion o supuesto del agente\n\n"
            "NO se escribio ningun hallazgo."
        )


def desglose_confianza(hallazgos):
    conteo = {nivel: 0 for nivel in NIVELES_CONFIANZA}
    for h in hallazgos:
        conteo[h["nivel_confianza"]] += 1
    return "Nivel de confianza: " + "  |  ".join(
        f"{nivel}: {conteo[nivel]}" for nivel in NIVELES_CONFIANZA
    )


def abrir_hoja(ruta_xlsx):
    wb = load_workbook(ruta_xlsx)
    if HOJA not in wb.sheetnames:
        raise ErrorEntrada(
            f"El archivo no tiene la hoja '{HOJA}'.\n"
            f"  Ruta: {os.path.abspath(ruta_xlsx)}\n"
            f"  Hojas encontradas: {', '.join(wb.sheetnames)}\n"
            "Usa un Excel generado desde la plantilla del plugin."
        )
    ws = wb[HOJA]
    verificar_esquema(ws, ruta_xlsx)
    return wb, ws


def ultima_fila_con_datos(ws):
    """Ultima fila con un hallazgo real (col C).

    El formato trae filas de plantilla con el ID pre-numerado pero vacias; por
    eso la ocupacion se mide por la columna 'hallazgo', no por el ID.
    """
    ultima = FILA_INICIO_DATOS - 1
    for fila in range(FILA_INICIO_DATOS, ws.max_row + 1):
        if ws.cell(fila, 3).value not in (None, ""):
            ultima = fila
    return ultima


def max_id(ws, hasta_fila):
    """Maximo ID entre filas que tienen un hallazgo real (ignora IDs de plantilla)."""
    maximo = 0
    for fila in range(FILA_INICIO_DATOS, hasta_fila + 1):
        if ws.cell(fila, 3).value in (None, ""):
            continue
        val = ws.cell(fila, 1).value
        try:
            maximo = max(maximo, int(val))
        except (TypeError, ValueError):
            continue
    return maximo


def guardar_atomico(wb, ruta_xlsx):
    """Guarda en un temporal y reemplaza el destino, para no dejarlo truncado."""
    destino = os.path.abspath(ruta_xlsx)
    carpeta = os.path.dirname(destino)
    fd, tmp = tempfile.mkstemp(suffix=".xlsx", prefix=".tmp_hallazgos_", dir=carpeta)
    os.close(fd)
    try:
        wb.save(tmp)
        os.replace(tmp, destino)
    except BaseException:
        try:
            os.remove(tmp)
        except OSError:
            pass
        raise


def cargar_hallazgos(ruta_json):
    try:
        with open(ruta_json, encoding="utf-8") as f:
            hallazgos = json.load(f)
    except FileNotFoundError:
        raise ErrorEntrada(f"No se encontro el JSON de hallazgos: {ruta_json}")
    except json.JSONDecodeError as err:
        raise ErrorEntrada(f"El JSON de hallazgos no es valido ({ruta_json}): {err}")
    if not isinstance(hallazgos, list):
        raise ErrorEntrada(
            f"El JSON debe ser una LISTA de hallazgos; se recibio {type(hallazgos).__name__}."
        )
    return hallazgos


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("json")
    ap.add_argument("xlsx", help="Excel destino; se crea desde la plantilla si no existe")
    ap.add_argument("--fecha", default=dt.date.today().isoformat())
    ap.add_argument("--dry-run", action="store_true", help="No guarda; solo reporta que haria")
    args = ap.parse_args()

    try:
        hallazgos = cargar_hallazgos(args.json)
        validar_hallazgos(hallazgos)
        creado = asegurar_destino(args.xlsx, dry_run=args.dry_run)
    except ErrorEntrada as err:
        sys.stderr.write(f"ERROR: {err}\n")
        return SALIDA_ENTRADA
    except PermissionError as err:
        sys.stderr.write(mensaje_permisos(args.xlsx, err))
        return SALIDA_PERMISOS
    except OSError as err:
        sys.stderr.write(
            f"ERROR: no se pudo crear el Excel destino.\n"
            f"  Ruta: {os.path.abspath(args.xlsx)}\n"
            f"  Detalle: {err}\n"
        )
        return SALIDA_PERMISOS

    if creado and args.dry_run:
        print(f"DRY-RUN: se crearia desde la plantilla: {os.path.abspath(args.xlsx)}")
        print(f"Hallazgos a escribir: {len(hallazgos)}")
        print(desglose_confianza(hallazgos))
        print(f"Primera fila nueva: {FILA_INICIO_DATOS}  |  ID inicial: 1")
        print("DRY-RUN: no se guardo el archivo.")
        return SALIDA_OK

    if creado:
        print(f"Creado desde plantilla: {os.path.abspath(args.xlsx)}")

    try:
        wb, ws = abrir_hoja(args.xlsx)
    except ErrorEntrada as err:
        sys.stderr.write(f"ERROR: {err}\n")
        return SALIDA_ENTRADA
    except PermissionError as err:
        sys.stderr.write(mensaje_permisos(args.xlsx, err))
        return SALIDA_PERMISOS

    ultima = ultima_fila_con_datos(ws)
    siguiente_id = max_id(ws, ultima) + 1
    fila = ultima + 1

    for i, h in enumerate(hallazgos):
        ws.cell(fila, 1, siguiente_id + i)
        for col, clave in CAMPOS.items():
            ws.cell(fila, col, h.get(clave, ""))
        ws.cell(fila, COL_PLANO, h.get("plano_referencia", ""))
        ws.cell(fila, COL_FECHA, h.get("fecha_deteccion", args.fecha))
        # 'estado' es seguimiento manual del usuario: las filas nuevas siempre
        # nacen en 'Pendiente', se ignora cualquier valor que traiga el JSON.
        ws.cell(fila, COL_ESTADO, ESTADO_INICIAL)
        fila += 1

    resumen = (
        f"Hallazgos a escribir: {len(hallazgos)}\n"
        f"{desglose_confianza(hallazgos)}\n"
        f"Primera fila nueva: {ultima + 1}  |  ID inicial: {siguiente_id}"
    )
    if args.dry_run:
        print(resumen)
        print("DRY-RUN: no se guardo el archivo.")
        return SALIDA_OK

    try:
        guardar_atomico(wb, args.xlsx)
    except PermissionError as err:
        sys.stderr.write(mensaje_permisos(args.xlsx, err))
        return SALIDA_PERMISOS
    except OSError as err:
        sys.stderr.write(mensaje_permisos(args.xlsx, err))
        return SALIDA_PERMISOS

    print(resumen)
    print(f"OK: guardado en {os.path.abspath(args.xlsx)}")
    return SALIDA_OK


if __name__ == "__main__":
    sys.exit(main())
