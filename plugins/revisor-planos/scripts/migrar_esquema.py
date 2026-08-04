"""Migra un Excel de hallazgos del esquema v1 al v2, conservando los datos.

Esquema v1 -> v2:
    - se elimina la columna 'requiere_validacion'
    - se agrega 'nivel_confianza' (Alta/Media/Baja) justo despues de 'severidad'
    - 'estado' pasa de la columna K a la M (ultima), con su valor intacto

Como 'requiere_validacion' no equivale a un nivel de confianza, la conversion es
conservadora y nunca inventa un valor:
    requiere_validacion = TRUE   -> nivel_confianza = 'Baja'
    requiere_validacion = FALSE  -> nivel_confianza queda VACIO (lo completas tu)

Antes de escribir se guarda una copia de seguridad junto al original
(<archivo>.v1.bak.xlsx).

Uso:
    python migrar_esquema.py <ruta_xlsx> [--dry-run]

Codigos de salida:
    0  OK (o nada que hacer: el archivo ya estaba en v2)
    2  error de permisos / archivo en uso
    3  error de entrada (hoja ausente, esquema no reconocido)
    4  falta una dependencia de Python
"""
import argparse
import os
import shutil
import sys

from escribir_hallazgos import (
    COL_ESTADO,
    ESQUEMA,
    ESQUEMA_V1,
    FILA_ENCABEZADO,
    FILA_INICIO_DATOS,
    HOJA,
    NIVELES_CONFIANZA,
    SALIDA_ENTRADA,
    SALIDA_OK,
    SALIDA_PERMISOS,
    ErrorEntrada,
    excel_abierto,
    guardar_atomico,
    mensaje_permisos,
    ultima_fila_con_datos,
)

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:  # pragma: no cover - depende del entorno
    sys.stderr.write(
        "ERROR: falta la dependencia 'openpyxl'.\n"
        "Instalala con:\n"
        f'  "{sys.executable}" -m pip install openpyxl\n'
    )
    sys.exit(4)

ANCHOS = {
    "A": 6, "B": 16, "C": 30, "D": 20, "E": 12, "F": 16, "G": 24,
    "H": 26, "I": 34, "J": 30, "K": 22, "L": 16, "M": 14,
}
NOTA = (
    "Columnas B-J las escribe automaticamente el agente y K-L el script "
    "(no cambiar el orden). La columna M (estado) es de uso manual para hacer "
    "seguimiento a la resolucion."
)


def leer_encabezado(ws):
    valores = [ws.cell(FILA_ENCABEZADO, c).value for c in range(1, len(ESQUEMA) + 1)]
    return [("" if v is None else str(v).strip()) for v in valores]


def confianza_desde_validacion(valor):
    """TRUE -> 'Baja'; cualquier otra cosa -> vacio (no se inventa un nivel)."""
    if valor is None:
        return ""
    return "Baja" if str(valor).strip().lower() in ("true", "1", "si", "sí", "verdadero") else ""


def migrar_filas(ws, ultima):
    """Reescribe las filas de datos en el orden v2. Devuelve (migradas, sin_confianza)."""
    migradas = 0
    sin_confianza = 0
    for fila in range(FILA_INICIO_DATOS, ultima + 1):
        # Se lee TODA la fila v1 antes de escribir nada, para no pisar valores.
        v1 = {
            nombre: ws.cell(fila, col).value
            for col, nombre in enumerate(ESQUEMA_V1, start=1)
        }
        if v1["hallazgo"] in (None, ""):
            continue  # fila de plantilla vacia (solo trae el ID pre-numerado)

        nivel = confianza_desde_validacion(v1["requiere_validacion"])
        if not nivel:
            sin_confianza += 1

        nuevo = dict(v1)
        nuevo["nivel_confianza"] = nivel
        for col, nombre in enumerate(ESQUEMA, start=1):
            ws.cell(fila, col, nuevo.get(nombre, ""))
        migradas += 1

    # Las filas de plantilla vacias pueden arrastrar 'Pendiente' en la vieja K.
    for fila in range(FILA_INICIO_DATOS, ws.max_row + 1):
        if ws.cell(fila, 3).value in (None, "") and fila > ultima:
            ws.cell(fila, 11, None)
            ws.cell(fila, COL_ESTADO, None)
    return migradas, sin_confianza


def reconstruir_formato(ws):
    ws.cell(2, 1).value = NOTA
    for i, nombre in enumerate(ESQUEMA, start=1):
        ws.cell(FILA_ENCABEZADO, i).value = nombre
    for letra, ancho in ANCHOS.items():
        ws.column_dimensions[letra].width = ancho

    fin = max(40, ws.max_row)
    ws.data_validations.dataValidation = []

    def dv(col, opciones):
        v = DataValidation(
            type="list", formula1='"' + ",".join(opciones) + '"', allow_blank=True
        )
        ws.add_data_validation(v)
        v.add(f"{col}{FILA_INICIO_DATOS}:{col}{fin}")

    dv("B", ["Hidrosanitario", "Estructural", "Arquitectonico", "Electrico", "Otro"])
    dv("E", ["Crítica", "Alta", "Media", "Baja"])
    dv("F", list(NIVELES_CONFIANZA))
    dv("M", ["Pendiente", "En revisión", "Resuelto", "No aplica"])


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="Excel de hallazgos con el esquema anterior (v1)")
    ap.add_argument("--dry-run", action="store_true", help="No guarda; solo reporta que haria")
    args = ap.parse_args()

    ruta = os.path.abspath(args.xlsx)
    try:
        if not os.path.exists(ruta):
            raise ErrorEntrada(f"No se encontro el Excel: {ruta}")
        if excel_abierto(ruta) and not args.dry_run:
            raise ErrorEntrada(
                "El archivo esta ABIERTO en Excel (existe su archivo de bloqueo '~$').\n"
                f"  Ruta: {ruta}\n"
                "Cierralo y vuelve a ejecutar la migracion."
            )

        wb = load_workbook(ruta)
        if HOJA not in wb.sheetnames:
            raise ErrorEntrada(
                f"El archivo no tiene la hoja '{HOJA}'.\n"
                f"  Ruta: {ruta}\n"
                f"  Hojas encontradas: {', '.join(wb.sheetnames)}"
            )
        ws = wb[HOJA]
        encabezado = leer_encabezado(ws)

        if encabezado == ESQUEMA:
            print(f"El archivo ya usa el esquema actual (v2); no hay nada que migrar.\n  {ruta}")
            return SALIDA_OK
        if encabezado != ESQUEMA_V1:
            raise ErrorEntrada(
                "El encabezado (fila 4) no corresponde al esquema v1 ni al v2, asi que\n"
                "no se puede migrar de forma segura.\n"
                f"  Ruta: {ruta}\n"
                f"  Encontrado: {', '.join(encabezado)}\n"
                f"  Esperado (v1): {', '.join(ESQUEMA_V1)}\n"
                "Revisa el archivo manualmente."
            )
    except ErrorEntrada as err:
        sys.stderr.write(f"ERROR: {err}\n")
        return SALIDA_ENTRADA
    except PermissionError as err:
        sys.stderr.write(mensaje_permisos(ruta, err))
        return SALIDA_PERMISOS

    ultima = ultima_fila_con_datos(ws)
    migradas, sin_confianza = migrar_filas(ws, ultima)
    reconstruir_formato(ws)

    resumen = [
        f"Archivo: {ruta}",
        f"Hallazgos migrados de v1 a v2: {migradas}",
        f"  con nivel_confianza='Baja' (venian con requiere_validacion=TRUE): "
        f"{migradas - sin_confianza}",
        f"  con nivel_confianza VACIO (completalo a mano en la columna F): {sin_confianza}",
    ]
    if args.dry_run:
        print("\n".join(resumen))
        print("DRY-RUN: no se guardo el archivo.")
        return SALIDA_OK

    respaldo = ruta.rsplit(".xlsx", 1)[0] + ".v1.bak.xlsx"
    try:
        shutil.copyfile(ruta, respaldo)
        guardar_atomico(wb, ruta)
    except PermissionError as err:
        sys.stderr.write(mensaje_permisos(ruta, err))
        return SALIDA_PERMISOS
    except OSError as err:
        sys.stderr.write(mensaje_permisos(ruta, err))
        return SALIDA_PERMISOS

    print("\n".join(resumen))
    print(f"Copia de seguridad del original: {respaldo}")
    print("OK: migrado al esquema v2.")
    return SALIDA_OK


if __name__ == "__main__":
    sys.exit(main())
