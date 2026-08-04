"""Genera la plantilla templates/hallazgos_estructural.xlsx desde cero.

La plantilla no se versiona a mano: este script es la unica fuente de verdad de
su formato (titulo, nota, encabezado, anchos, alturas y listas desplegables).
Si el esquema cambia, se cambia 'ESQUEMA' en escribir_hallazgos.py y se vuelve a
correr este script; asi el encabezado real y el que valida el escritor no pueden
divergir.

Uso:
    python crear_plantilla.py [--salida <ruta_xlsx>] [--dry-run]

Por defecto escribe en templates/hallazgos_estructural.xlsx del plugin. Si el
archivo ya existe se sobrescribe, asi que no lo uses apuntando a un Excel de
hallazgos real: para eso esta escribir_hallazgos.py, que hace append.

Codigos de salida:
    0  OK
    2  error de permisos / archivo en uso
    4  falta una dependencia de Python
"""
import argparse
import os
import sys

from escribir_hallazgos import (
    ESQUEMA,
    FILA_ENCABEZADO,
    FILA_INICIO_DATOS,
    HOJA,
    NIVELES_CONFIANZA,
    PLANTILLA,
    SALIDA_OK,
    SALIDA_PERMISOS,
    guardar_atomico,
    mensaje_permisos,
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:  # pragma: no cover - depende del entorno
    sys.stderr.write(
        "ERROR: falta la dependencia 'openpyxl'.\n"
        "Instalala con:\n"
        f'  "{sys.executable}" -m pip install openpyxl\n'
    )
    sys.exit(4)

TITULO = "Tracker de hallazgos – Agente revisor estructural (NSR-10)"
NOTA = (
    "Columnas B-J las escribe automaticamente el agente y K-L el script "
    "(no cambiar el orden). La columna M (estado) es de uso manual para hacer "
    "seguimiento a la resolucion."
)

AZUL = "FF1F4E78"
GRIS_NOTA = "FF595959"
GRIS_ID = "FF808080"
BLANCO = "FFFFFFFF"

ANCHOS = {
    "A": 6, "B": 16, "C": 30, "D": 20, "E": 12, "F": 16, "G": 24,
    "H": 26, "I": 34, "J": 30, "K": 22, "L": 16, "M": 14,
}

ALTO_TITULO = 21.75
ALTO_NOTA = 15.75
ALTO_ENCABEZADO = 27.6
ALTO_DATOS = 39.75

# Hasta donde se preparan formato y validaciones. No es un tope de filas: al
# agregar hallazgos mas alla de la 40 el script sigue escribiendo, solo que esas
# filas nacen sin desplegable.
FILA_FIN = 40
IDS_PRENUMERADOS = 4

# La lista de disciplinas incluye las demas a proposito: permite consolidar en un
# solo archivo los hallazgos de este plugin y los del revisor hidrosanitario.
DISCIPLINAS = ["Estructural", "Hidrosanitario", "Arquitectonico", "Electrico", "Otro"]
SEVERIDADES = ["Crítica", "Alta", "Media", "Baja"]
ESTADOS = ["Pendiente", "En revisión", "Resuelto", "No aplica"]


def construir():
    wb = Workbook()
    ws = wb.active
    ws.title = HOJA
    ultima_col = len(ESQUEMA)
    letra_fin = chr(64 + ultima_col)

    ws.merge_cells(f"A1:{letra_fin}1")
    celda_titulo = ws.cell(1, 1, TITULO)
    celda_titulo.font = Font(bold=True, size=14, color=AZUL)
    ws.row_dimensions[1].height = ALTO_TITULO

    ws.merge_cells(f"A2:{letra_fin}2")
    celda_nota = ws.cell(2, 1, NOTA)
    celda_nota.font = Font(size=10, color=GRIS_NOTA)
    ws.row_dimensions[2].height = ALTO_NOTA

    # La fila 3 queda vacia a proposito: separa la cabecera del encabezado.

    relleno = PatternFill("solid", fgColor=AZUL)
    for i, nombre in enumerate(ESQUEMA, start=1):
        celda = ws.cell(FILA_ENCABEZADO, i, nombre)
        celda.font = Font(bold=True, color=BLANCO)
        celda.fill = relleno
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[FILA_ENCABEZADO].height = ALTO_ENCABEZADO

    for fila in range(FILA_INICIO_DATOS, FILA_FIN + 1):
        ws.row_dimensions[fila].height = ALTO_DATOS
        for col in range(1, ultima_col + 1):
            ws.cell(fila, col).alignment = Alignment(vertical="top", wrap_text=True)

    # IDs de ejemplo en gris. Por esto ultima_fila_con_datos() mide la ocupacion
    # por la columna 'hallazgo' y no por el ID: estas filas estan vacias.
    for n in range(1, IDS_PRENUMERADOS + 1):
        celda = ws.cell(FILA_INICIO_DATOS + n - 1, 1, n)
        celda.font = Font(color=GRIS_ID)

    for letra, ancho in ANCHOS.items():
        ws.column_dimensions[letra].width = ancho

    ws.freeze_panes = f"A{FILA_INICIO_DATOS}"

    def desplegable(col, opciones):
        v = DataValidation(
            type="list", formula1='"' + ",".join(opciones) + '"', allow_blank=True
        )
        ws.add_data_validation(v)
        v.add(f"{col}{FILA_INICIO_DATOS}:{col}{FILA_FIN}")

    desplegable("B", DISCIPLINAS)
    desplegable("E", SEVERIDADES)
    desplegable("F", list(NIVELES_CONFIANZA))
    desplegable("M", ESTADOS)

    return wb


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--salida", default=PLANTILLA, help="Ruta del .xlsx a generar")
    ap.add_argument("--dry-run", action="store_true", help="No guarda; solo reporta que haria")
    args = ap.parse_args()

    destino = os.path.abspath(args.salida)
    existia = os.path.exists(destino)

    if args.dry_run:
        print(f"DRY-RUN: se generaria la plantilla en {destino}")
        print(f"  Hoja: {HOJA}")
        print(f"  Columnas ({len(ESQUEMA)}): {', '.join(ESQUEMA)}")
        print(f"  {'SOBRESCRIBIRIA el archivo existente' if existia else 'Archivo nuevo'}")
        return SALIDA_OK

    wb = construir()
    try:
        os.makedirs(os.path.dirname(destino), exist_ok=True)
        if not existia:
            # guardar_atomico necesita que la carpeta exista, pero no el archivo.
            wb.save(destino)
        else:
            guardar_atomico(wb, destino)
    except PermissionError as err:
        sys.stderr.write(mensaje_permisos(destino, err))
        return SALIDA_PERMISOS
    except OSError as err:
        sys.stderr.write(mensaje_permisos(destino, err))
        return SALIDA_PERMISOS

    print(f"OK: plantilla {'regenerada' if existia else 'creada'} en {destino}")
    print(f"  Hoja: {HOJA}  |  Columnas: {len(ESQUEMA)}")
    return SALIDA_OK


if __name__ == "__main__":
    sys.exit(main())
