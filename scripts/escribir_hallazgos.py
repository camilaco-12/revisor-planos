"""Escribe hallazgos hidrosanitarios en el Excel del proyecto (append).

Lee un JSON (lista de objetos con las 9 claves del esquema + 'plano_referencia')
y agrega una fila por hallazgo en la hoja 'Hallazgos hidrosanitarios' del Excel
destino, sin sobrescribir filas existentes.

Si el Excel destino no existe, se crea automaticamente copiando la plantilla
templates/hallazgos_hidrosanitario.xlsx del plugin.

Mapeo de columnas (encabezado en la fila 4, datos desde la fila 5):
    A  ID                     -> consecutivo, continua el maximo existente
    B  disciplina
    C  hallazgo
    D  ubicacion
    E  severidad
    F  referencia_normativa
    G  riesgo
    H  justificacion_tecnica
    I  accion_correctiva
    J  requiere_validacion    -> TRUE / FALSE
    K  estado                 -> por defecto 'Pendiente'
    L  plano_referencia
    M  fecha_deteccion        -> por defecto la fecha de hoy

Uso:
    python escribir_hallazgos.py <ruta_json> <ruta_xlsx> [--fecha AAAA-MM-DD] [--dry-run]

Codigos de salida:
    0  OK
    2  error de permisos / archivo en uso (mensaje accionable en stderr)
    3  error de entrada (JSON invalido, plantilla ausente, hoja inexistente)
    4  falta una dependencia de Python
"""
import argparse
import datetime as dt
import json
import os
import shutil
import sys
import tempfile

SALIDA_OK = 0
SALIDA_PERMISOS = 2
SALIDA_ENTRADA = 3
SALIDA_DEPENDENCIA = 4

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - depende del entorno
    sys.stderr.write(
        "ERROR: falta la dependencia 'openpyxl'.\n"
        "Instalala con:\n"
        f'  "{sys.executable}" -m pip install openpyxl\n'
    )
    sys.exit(SALIDA_DEPENDENCIA)

HOJA = "Hallazgos hidrosanitarios"
FILA_ENCABEZADO = 4
FILA_INICIO_DATOS = 5
DIR = os.path.dirname(os.path.abspath(__file__))
RAIZ_PLUGIN = os.environ.get("CLAUDE_PLUGIN_ROOT") or os.path.dirname(DIR)
PLANTILLA = os.path.join(RAIZ_PLUGIN, "templates", "hallazgos_hidrosanitario.xlsx")

# columna (indice 1-based) -> clave del JSON
CAMPOS = {
    2: "disciplina",
    3: "hallazgo",
    4: "ubicacion",
    5: "severidad",
    6: "referencia_normativa",
    7: "riesgo",
    8: "justificacion_tecnica",
    9: "accion_correctiva",
}


class ErrorEntrada(Exception):
    """Problema con los datos o rutas de entrada; mensaje ya legible."""


def bool_txt(valor):
    if isinstance(valor, bool):
        return "TRUE" if valor else "FALSE"
    return "TRUE" if str(valor).strip().lower() in ("true", "1", "si", "sí") else "FALSE"


def excel_abierto(ruta):
    """Detecta el lockfile '~$archivo.xlsx' que Excel crea al abrir un libro."""
    carpeta, nombre = os.path.split(os.path.abspath(ruta))
    return os.path.exists(os.path.join(carpeta, "~$" + nombre))


def mensaje_permisos(ruta, err):
    """Mensaje accionable para un fallo de escritura en Windows."""
    lineas = [
        "ERROR DE PERMISOS: no se pudo escribir el archivo.",
        f"  Ruta: {os.path.abspath(ruta)}",
        f"  Detalle del sistema: {err}",
        "",
    ]
    if excel_abierto(ruta):
        lineas += [
            "Causa mas probable: el archivo esta ABIERTO en Excel.",
            "Como solucionarlo:",
            "  1. Cierra el archivo en Excel (incluida cualquier ventana de solo lectura).",
            "  2. Vuelve a ejecutar el comando.",
        ]
    else:
        lineas += [
            "Causas posibles (en orden de probabilidad):",
            "  1. El archivo esta abierto en Excel -> cierralo y reintenta.",
            "  2. Acceso controlado a carpetas de Windows Defender esta bloqueando la",
            "     escritura (tipico en Escritorio, Documentos, Imagenes). Ve a",
            "     Seguridad de Windows -> Proteccion contra virus y amenazas ->",
            "     Proteccion contra ransomware -> Permitir una app a traves del acceso",
            "     controlado, y autoriza:",
            f"       {sys.executable}",
            "  3. La carpeta es de solo lectura o esta en OneDrive sin sincronizar ->",
            "     elige otra ruta destino, por ejemplo dentro de la carpeta del proyecto.",
        ]
    lineas += [
        "",
        "NO se escribio ningun hallazgo. El archivo destino quedo intacto.",
    ]
    return "\n".join(lineas) + "\n"


def asegurar_destino(ruta_xlsx, dry_run=False):
    """Crea el Excel destino desde la plantilla si aun no existe.

    Devuelve True si lo creo (o lo crearia, en dry-run).
    """
    if os.path.exists(ruta_xlsx):
        return False

    if not os.path.exists(PLANTILLA):
        raise ErrorEntrada(
            "No existe el Excel destino y tampoco se encontro la plantilla del plugin.\n"
            f"  Destino: {os.path.abspath(ruta_xlsx)}\n"
            f"  Plantilla esperada: {PLANTILLA}\n"
            "Verifica la instalacion del plugin o crea el Excel manualmente."
        )

    if dry_run:
        return True

    carpeta = os.path.dirname(os.path.abspath(ruta_xlsx))
    os.makedirs(carpeta, exist_ok=True)
    shutil.copyfile(PLANTILLA, ruta_xlsx)
    return True


def abrir_hoja(ruta_xlsx):
    wb = load_workbook(ruta_xlsx)
    if HOJA not in wb.sheetnames:
        raise ErrorEntrada(
            f"El archivo no tiene la hoja '{HOJA}'.\n"
            f"  Ruta: {os.path.abspath(ruta_xlsx)}\n"
            f"  Hojas encontradas: {', '.join(wb.sheetnames)}\n"
            "Usa un Excel generado desde la plantilla del plugin."
        )
    return wb, wb[HOJA]


def ultima_fila_con_datos(ws):
    """Ultima fila con un hallazgo real (col C).

    El formato trae filas de plantilla con el ID pre-numerado pero vacias; por
    eso la ocupacion se mide por la columna 'hallazgo', no por el ID.
    """
    ultima = FILA_INICIO_DATOS - 1
    for fila in range(FILA_INICIO_DATOS, ws.max_row + 1):
        if ws.cell(fila, 3).value not in (None, ""):
            ultima = fila
    return ultima


def max_id(ws, hasta_fila):
    """Maximo ID entre filas que tienen un hallazgo real (ignora IDs de plantilla)."""
    maximo = 0
    for fila in range(FILA_INICIO_DATOS, hasta_fila + 1):
        if ws.cell(fila, 3).value in (None, ""):
            continue
        val = ws.cell(fila, 1).value
        try:
            maximo = max(maximo, int(val))
        except (TypeError, ValueError):
            continue
    return maximo


def guardar_atomico(wb, ruta_xlsx):
    """Guarda en un temporal y reemplaza el destino, para no dejarlo truncado."""
    destino = os.path.abspath(ruta_xlsx)
    carpeta = os.path.dirname(destino)
    fd, tmp = tempfile.mkstemp(suffix=".xlsx", prefix=".tmp_hallazgos_", dir=carpeta)
    os.close(fd)
    try:
        wb.save(tmp)
        os.replace(tmp, destino)
    except BaseException:
        try:
            os.remove(tmp)
        except OSError:
            pass
        raise


def cargar_hallazgos(ruta_json):
    try:
        with open(ruta_json, encoding="utf-8") as f:
            hallazgos = json.load(f)
    except FileNotFoundError:
        raise ErrorEntrada(f"No se encontro el JSON de hallazgos: {ruta_json}")
    except json.JSONDecodeError as err:
        raise ErrorEntrada(f"El JSON de hallazgos no es valido ({ruta_json}): {err}")
    if not isinstance(hallazgos, list):
        raise ErrorEntrada(
            f"El JSON debe ser una LISTA de hallazgos; se recibio {type(hallazgos).__name__}."
        )
    return hallazgos


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("json")
    ap.add_argument("xlsx", help="Excel destino; se crea desde la plantilla si no existe")
    ap.add_argument("--fecha", default=dt.date.today().isoformat())
    ap.add_argument("--dry-run", action="store_true", help="No guarda; solo reporta que haria")
    args = ap.parse_args()

    try:
        hallazgos = cargar_hallazgos(args.json)
        creado = asegurar_destino(args.xlsx, dry_run=args.dry_run)
    except ErrorEntrada as err:
        sys.stderr.write(f"ERROR: {err}\n")
        return SALIDA_ENTRADA
    except PermissionError as err:
        sys.stderr.write(mensaje_permisos(args.xlsx, err))
        return SALIDA_PERMISOS
    except OSError as err:
        sys.stderr.write(
            f"ERROR: no se pudo crear el Excel destino.\n"
            f"  Ruta: {os.path.abspath(args.xlsx)}\n"
            f"  Detalle: {err}\n"
        )
        return SALIDA_PERMISOS

    if creado and args.dry_run:
        print(f"DRY-RUN: se crearia desde la plantilla: {os.path.abspath(args.xlsx)}")
        print(f"Hallazgos a escribir: {len(hallazgos)}")
        print(f"Primera fila nueva: {FILA_INICIO_DATOS}  |  ID inicial: 1")
        print("DRY-RUN: no se guardo el archivo.")
        return SALIDA_OK

    if creado:
        print(f"Creado desde plantilla: {os.path.abspath(args.xlsx)}")

    try:
        wb, ws = abrir_hoja(args.xlsx)
    except ErrorEntrada as err:
        sys.stderr.write(f"ERROR: {err}\n")
        return SALIDA_ENTRADA
    except PermissionError as err:
        sys.stderr.write(mensaje_permisos(args.xlsx, err))
        return SALIDA_PERMISOS

    ultima = ultima_fila_con_datos(ws)
    siguiente_id = max_id(ws, ultima) + 1
    fila = ultima + 1

    for i, h in enumerate(hallazgos):
        ws.cell(fila, 1, siguiente_id + i)
        for col, clave in CAMPOS.items():
            ws.cell(fila, col, h.get(clave, ""))
        ws.cell(fila, 10, bool_txt(h.get("requiere_validacion", False)))
        ws.cell(fila, 11, h.get("estado", "Pendiente"))
        ws.cell(fila, 12, h.get("plano_referencia", ""))
        ws.cell(fila, 13, h.get("fecha_deteccion", args.fecha))
        fila += 1

    resumen = (
        f"Hallazgos a escribir: {len(hallazgos)}\n"
        f"Primera fila nueva: {ultima + 1}  |  ID inicial: {siguiente_id}"
    )
    if args.dry_run:
        print(resumen)
        print("DRY-RUN: no se guardo el archivo.")
        return SALIDA_OK

    try:
        guardar_atomico(wb, args.xlsx)
    except PermissionError as err:
        sys.stderr.write(mensaje_permisos(args.xlsx, err))
        return SALIDA_PERMISOS
    except OSError as err:
        sys.stderr.write(mensaje_permisos(args.xlsx, err))
        return SALIDA_PERMISOS

    print(resumen)
    print(f"OK: guardado en {os.path.abspath(args.xlsx)}")
    return SALIDA_OK


if __name__ == "__main__":
    sys.exit(main())
